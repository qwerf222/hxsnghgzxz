from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "副本公招角色25-12-17.xlsx"
OUTPUT_PATH = ROOT / "app/src/main/java/com/gzxz/recruit/GeneratedRecruitData.kt"
SHEET_NAME = "Sheet1 (2)"
DATA_START_ROW = 11


@dataclass(frozen=True)
class GroupDefinition:
    key: str
    label: str
    order: int
    base_priority: int


GROUP_DEFINITIONS = [
    GroupDefinition("profession", "职业标签", 0, 120),
    GroupDefinition("species", "种族标签", 1, 100),
    GroupDefinition("element", "属性标签", 2, 80),
    GroupDefinition("terrain", "地形标签", 3, 60),
]


def kotlin_string(value: str) -> str:
    escaped = value.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def parse_operators() -> list[dict]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook[SHEET_NAME]
    operators_by_name: dict[str, dict] = {}

    for row in sheet.iter_rows(min_row=DATA_START_ROW, values_only=True):
        blocks = [
            (row[0], row[1:5], row[6]),
            (row[7], row[8:12], row[12]),
        ]
        for name, raw_tags, raw_rarity in blocks:
            if not name:
                continue

            tags = tuple(tag for tag in raw_tags if tag)
            rarity = int(raw_rarity)
            record = {
                "name": str(name).strip(),
                "rarity": rarity,
                "tags": tags,
            }
            existing = operators_by_name.get(record["name"])
            if existing and existing != record:
                raise ValueError(f"角色 {record['name']} 在 Excel 中存在冲突数据: {existing} != {record}")
            operators_by_name[record["name"]] = record

    return sorted(
        operators_by_name.values(),
        key=lambda item: (-item["rarity"], item["name"]),
    )


def build_tag_entries(operators: list[dict]) -> list[dict]:
    stats_by_tag: dict[str, dict] = defaultdict(lambda: {"count": 0, "max_rarity": 0, "group": None})

    for operator in operators:
        for group, tag in zip(GROUP_DEFINITIONS, operator["tags"], strict=True):
            stats = stats_by_tag[tag]
            stats["count"] += 1
            stats["max_rarity"] = max(stats["max_rarity"], operator["rarity"])
            stats["group"] = group

    tags = []
    for tag_name, stats in stats_by_tag.items():
        group: GroupDefinition = stats["group"]
        priority = group.base_priority + stats["max_rarity"] * 25 + round(300 / stats["count"])
        tags.append(
            {
                "name": tag_name,
                "group_key": group.key,
                "priority": priority,
                "group_order": group.order,
            }
        )

    return sorted(tags, key=lambda item: (item["group_order"], -item["priority"], item["name"]))


def render_kotlin(operators: list[dict], tags: list[dict]) -> str:
    lines: list[str] = []
    lines.append("package com.gzxz.recruit")
    lines.append("")
    lines.append("/**")
    lines.append(" * Generated from `副本公招角色25-12-17.xlsx` by `tools/generate_recruit_data.py`.")
    lines.append(" * Do not edit manually; regenerate when the workbook changes.")
    lines.append(" */")
    lines.append("object GeneratedRecruitData {")
    lines.append("    val tagGroups: List<TagGroup> = listOf(")
    for group in GROUP_DEFINITIONS:
        lines.append(
            f"        TagGroup(key = {kotlin_string(group.key)}, label = {kotlin_string(group.label)}, order = {group.order}),"
        )
    lines.append("    )")
    lines.append("")
    lines.append("    val tags: List<RecruitTag> = listOf(")
    for tag in tags:
        lines.append(
            "        RecruitTag(" \
            f"name = {kotlin_string(tag['name'])}, " \
            f"groupKey = {kotlin_string(tag['group_key'])}, " \
            f"priority = {tag['priority']}" \
            "),"
        )
    lines.append("    )")
    lines.append("")
    lines.append("    val operators: List<RecruitOperator> = listOf(")
    for operator in operators:
        tag_list = ", ".join(kotlin_string(tag) for tag in operator["tags"])
        lines.append(
            "        RecruitOperator(" \
            f"name = {kotlin_string(operator['name'])}, " \
            f"rarity = {operator['rarity']}, " \
            f"tags = setOf({tag_list})" \
            "),"
        )
    lines.append("    )")
    lines.append("}")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    operators = parse_operators()
    tags = build_tag_entries(operators)
    OUTPUT_PATH.write_text(render_kotlin(operators, tags), encoding="utf-8")
    print(f"Generated {len(operators)} operators and {len(tags)} tags -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

